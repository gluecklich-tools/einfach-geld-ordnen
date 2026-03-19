from __future__ import annotations

import argparse
import json
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


NAVY = "1F497D"
NAVY_DARK = "17375E"
BLUE_LIGHT = "DCE8F7"
BLUE_PALE = "EEF4FB"
WHITE = "FFFFFF"
TEXT_DARK = "1F1F1F"
TEXT_MID = "44546A"
GRID = "C9D5E6"


def solid_fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def thin_border(color: str = GRID) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_range(
    ws: Worksheet,
    cell_range: str,
    *,
    fill: str | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = solid_fill(fill)
            if font is not None:
                cell.font = copy(font)
            if alignment is not None:
                cell.alignment = copy(alignment)
            if border is not None:
                cell.border = copy(border)


def apply_start_refine(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    ws.row_dimensions[1].height = 44
    ws.row_dimensions[2].height = 4
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 5
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 34
    ws.row_dimensions[7].height = 4
    ws.row_dimensions[8].height = 42
    ws.row_dimensions[9].height = 26
    ws.row_dimensions[10].height = 26
    ws.row_dimensions[11].height = 7
    ws.row_dimensions[12].height = 9
    ws.row_dimensions[13].height = 24
    ws.row_dimensions[14].height = 24
    ws.row_dimensions[15].height = 24
    ws.row_dimensions[16].height = 24
    ws.row_dimensions[17].height = 24
    ws.row_dimensions[18].height = 24
    ws.row_dimensions[19].height = 8
    ws.row_dimensions[20].height = 30

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 23
    ws.column_dimensions["E"].width = 23
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 3
    changes.append("START density/presence layout applied")

    style_range(
        ws,
        "A1:I1",
        fill=NAVY,
        font=Font(name="Calibri", size=24, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "A3:F3",
        fill="EAF1FA",
        font=Font(name="Calibri", size=10, bold=True, color=TEXT_MID),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "G3:I3",
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A5:C6",
        fill=BLUE_PALE,
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A8:C10",
        fill=BLUE_PALE,
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D5:F6",
        fill="FCFDFE",
        font=Font(name="Calibri", size=18, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D8:F8",
        fill="E6EFFB",
        font=Font(name="Calibri", size=21, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D9:F10",
        fill=BLUE_PALE,
        font=Font(name="Calibri", size=13, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "G5:I5",
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "G6:I10",
        fill="F2F7FC",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    style_range(
        ws,
        "A13:I13",
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=12, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A14:F14",
        fill="F8FBFE",
        font=Font(name="Calibri", size=11, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A15:F15",
        fill="F3F7FC",
        font=Font(name="Calibri", size=11, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A16:F16",
        fill="F8FBFE",
        font=Font(name="Calibri", size=11, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A17:F17",
        fill="F3F7FC",
        font=Font(name="Calibri", size=11, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A18:F18",
        fill="F8FBFE",
        font=Font(name="Calibri", size=11, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "G14:I18",
        fill="EDF4FB",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A20:I20",
        fill="F0F5FA",
        font=Font(name="Calibri", size=10, italic=True, color=TEXT_MID),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("START visual refine styling applied")
    return changes


def collect_snapshot(ws: Worksheet, *, max_row: int = 120, max_col: int = 24) -> dict[str, Any]:
    used_rows = 0
    used_cols = 0
    values: list[dict[str, Any]] = []

    row_cap = min(max_row, ws.max_row)
    col_cap = min(max_col, ws.max_column)

    for row in range(1, row_cap + 1):
        row_has_value = False
        for col in range(1, col_cap + 1):
            value = ws.cell(row=row, column=col).value
            if value not in (None, ""):
                row_has_value = True
                used_rows = max(used_rows, row)
                used_cols = max(used_cols, col)
                values.append(
                    {
                        "row": row,
                        "col": col,
                        "coord": ws.cell(row=row, column=col).coordinate,
                        "value": str(value),
                    }
                )
        if row_has_value:
            used_rows = max(used_rows, row)

    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]

    row_heights: dict[str, float] = {}
    for idx in range(1, min(max_row, 160) + 1):
        height = ws.row_dimensions[idx].height
        if height is not None:
            row_heights[str(idx)] = height

    col_widths: dict[str, float] = {}
    for key in [chr(code) for code in range(ord("A"), ord("X") + 1)]:
        width = ws.column_dimensions[key].width
        if width is not None:
            col_widths[key] = width

    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "used_rows_in_snapshot": used_rows,
        "used_cols_in_snapshot": used_cols,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "merged_ranges": merged_ranges,
        "row_heights": row_heights,
        "col_widths": col_widths,
        "values": values,
    }


def apply_haushaltsbuch_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Step3: final small readability lift without logic changes
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 25
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 15
    ws.row_dimensions[9].height = 22
    ws.row_dimensions[10].height = 14
    ws.row_dimensions[11].height = 22
    ws.row_dimensions[12].height = 14
    ws.row_dimensions[13].height = 20

    style_range(
        ws,
        "A3:H3",
        fill="D2E2F4",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A4:H8",
        fill="FFFFFF",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "E4:E8",
        fill="F4F8FD",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "T3:T8",
        fill="F2F6FB",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    style_range(
        ws,
        "T9:T13",
        fill="FAFCFE",
        font=Font(name="Calibri", size=9, italic=True, color=TEXT_MID),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )

    changes.append("final small readability lift applied to headers, first rows and help box")
    return changes



def apply_monat_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Step3: final width balance and stronger monthly focus without rebuild
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[3].height = 38
    ws.row_dimensions[5].height = 38
    ws.row_dimensions[6].height = 38
    ws.row_dimensions[7].height = 38
    ws.row_dimensions[8].height = 32
    ws.row_dimensions[9].height = 32
    ws.row_dimensions[11].height = 34
    for idx in range(12, 23):
        ws.row_dimensions[idx].height = 26

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 4
    ws.column_dimensions["I"].width = 4
    ws.column_dimensions["J"].width = 40
    changes.append("monthly block widened again for final balance")

    style_range(
        ws,
        "A3:C3",
        fill="D2E2F4",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D3:F3",
        fill="F4F8FC",
        font=Font(name="Calibri", size=12, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    changes.append("focus month row balanced for stronger monthly entry")

    style_range(
        ws,
        "A6:C9",
        fill="BDD2E8",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D6:F7",
        fill="FFFFFF",
        font=Font(name="Calibri", size=13, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D8:F8",
        fill="DEEAF7",
        font=Font(name="Calibri", size=16, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D9:F9",
        fill="F7FAFD",
        font=Font(name="Calibri", size=10, italic=True, color=TEXT_MID),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("overview KPIs given final readability lift")

    style_range(
        ws,
        "A12:F12",
        fill="D2E2F4",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A13:B22",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "C13:E22",
        fill="FFFFFF",
        font=Font(name="Calibri", size=10, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "F13:F22",
        fill="F1F6FB",
        font=Font(name="Calibri", size=10, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("category table made easier to scan in final pass")

    style_range(
        ws,
        "J2:J7",
        fill="F7FAFD",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    changes.append("info box kept soft and visually secondary")

    return changes

def apply_budgets_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Step2: final scanability and calmer top planning block
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[5].height = 32
    ws.row_dimensions[6].height = 32
    ws.row_dimensions[7].height = 32
    ws.row_dimensions[11].height = 26
    for idx in range(12, 37):
        ws.row_dimensions[idx].height = 23

    ws.column_dimensions["A"].width = 27
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["J"].width = 41
    changes.append("budget columns widened slightly for final readability")

    style_range(
        ws,
        "A3:A3",
        fill="D6E5F6",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B3:C3",
        fill="F2F6FB",
        font=Font(name="Calibri", size=12, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    changes.append("focus row balanced")

    style_range(
        ws,
        "A5:A7",
        fill="35689A",
        font=Font(name="Calibri", size=10, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "B5:B7",
        fill="FFF8E2",
        font=Font(name="Calibri", size=12, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "C5:F7",
        fill="FCFAEF",
        font=Font(name="Calibri", size=10, color=TEXT_MID),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("planning block softened slightly")

    style_range(
        ws,
        "A11:F11",
        fill="35689A",
        font=Font(name="Calibri", size=11, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "A12:A36",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B12:E36",
        fill="FFFFFF",
        font=Font(name="Calibri", size=10, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "F12:F36",
        fill="F1F6FB",
        font=Font(name="Calibri", size=10, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("table scanability improved in final pass")

    style_range(
        ws,
        "J2:J7",
        fill="F6F9FD",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    changes.append("info box kept soft and secondary")

    return changes

def apply_fixkosten_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Step2: final width balance and scanability for fixed costs
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 31
    ws.row_dimensions[5].height = 31
    ws.row_dimensions[6].height = 31
    ws.row_dimensions[7].height = 31
    for idx in range(8, 31):
        ws.row_dimensions[idx].height = 23

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 17
    ws.column_dimensions["J"].width = 42
    changes.append("fixkosten table widened slightly for final scan balance")

    style_range(
        ws,
        "A3:C3",
        fill="35689A",
        font=Font(name="Calibri", size=11, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "D3:F3",
        fill="F2F6FB",
        font=Font(name="Calibri", size=13, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    changes.append("summary block balanced in final pass")

    style_range(
        ws,
        "A4:A30",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B4:B30",
        fill="FFF9E6",
        font=Font(name="Calibri", size=10, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "C4:E30",
        fill="FCFDFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "F4:F30",
        fill="F1F6FB",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    changes.append("table columns tuned for calmer row scan")

    style_range(
        ws,
        "J2:J6",
        fill="F6F9FD",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    changes.append("info box kept soft and secondary")

    return changes

def apply_planung_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Step1: stronger planning controls and calmer 12-month table
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 4
    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 4
    ws.row_dimensions[5].height = 30
    for idx in range(6, 18):
        ws.row_dimensions[idx].height = 24

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 3
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 40
    changes.append("planung layout widened for control block and table readability")

    style_range(
        ws,
        "A1:I1",
        fill=NAVY,
        font=Font(name="Calibri", size=16, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "J1:J1",
        fill=NAVY,
        font=Font(name="Calibri", size=11, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )

    style_range(
        ws,
        "A3:A3",
        fill="DCE8F7",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B3:B3",
        fill="EEF4FB",
        font=Font(name="Calibri", size=12, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D3:E3",
        fill="DCE8F7",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    changes.append("control row made clearer and more stable")

    style_range(
        ws,
        "A5:G5",
        fill="35689A",
        font=Font(name="Calibri", size=10, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "A6:A17",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B6:F17",
        fill="FFFFFF",
        font=Font(name="Calibri", size=10, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "G6:G17",
        fill="EEF4FB",
        font=Font(name="Calibri", size=10, color=TEXT_MID),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("12-month planning table made easier to scan")

    style_range(
        ws,
        "J2:J7",
        fill="F3F7FC",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    changes.append("planning info box softened and integrated")

    return changes

def apply_jahr_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Step1: stronger year header and calmer annual table
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 4
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[4].height = 4
    ws.row_dimensions[5].height = 30
    for idx in range(6, 18):
        ws.row_dimensions[idx].height = 24

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 40
    changes.append("jahr layout widened for header and table readability")

    style_range(
        ws,
        "A1:I1",
        fill=NAVY,
        font=Font(name="Calibri", size=16, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "J1:J1",
        fill=NAVY,
        font=Font(name="Calibri", size=11, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )

    style_range(
        ws,
        "A3:A3",
        fill="DCE8F7",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B3:B3",
        fill="EEF4FB",
        font=Font(name="Calibri", size=13, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    changes.append("year focus block made clearer")

    style_range(
        ws,
        "A5:F5",
        fill="35689A",
        font=Font(name="Calibri", size=10, bold=True, color=WHITE),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "A6:B17",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "C6:F17",
        fill="FFFFFF",
        font=Font(name="Calibri", size=10, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="right", vertical="center"),
        border=thin_border(),
    )
    changes.append("annual table made easier to scan")

    style_range(
        ws,
        "J3:J7",
        fill="F3F7FC",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        border=thin_border(),
    )
    changes.append("year info box softened and integrated")

    return changes

def apply_notgroschen_scaffold(ws: Worksheet) -> list[str]:
    changes: list[str] = []

    # Structural fix: left analysis block + real separated right info panel
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 4
    for idx in range(3, 8):
        ws.row_dimensions[idx].height = 34
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 24
    for idx in range(10, 19):
        ws.row_dimensions[idx].height = 24

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 4
    ws.column_dimensions["I"].width = 4
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14
    changes.append("notgroschen layout rebuilt with real right-side panel separation")

    for merge_range in ["J1:J1", "J2:J7", "J1:L1", "J2:L7", "J1:M1", "J2:M7"]:
        try:
            ws.unmerge_cells(merge_range)
        except Exception:
            pass

    ws.merge_cells("J1:M1")
    ws.merge_cells("J2:M7")
    changes.append("notgroschen info panel rebuilt across J:M")

    style_range(
        ws,
        "A1:G1",
        fill=NAVY,
        font=Font(name="Calibri", size=16, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )
    style_range(
        ws,
        "J1:M1",
        fill=NAVY,
        font=Font(name="Calibri", size=11, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )

    style_range(
        ws,
        "A3:A7",
        fill="AFC8E2",
        font=Font(name="Calibri", size=11, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B3:B7",
        fill="F6F1D8",
        font=Font(name="Calibri", size=12, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "D3:D7",
        fill="F7F9FC",
        font=Font(name="Calibri", size=11, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "E3:E7",
        fill="F6F1D8",
        font=Font(name="Calibri", size=12, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "F3:G7",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("top overview block compacted and stabilized")

    style_range(
        ws,
        "A9:G9",
        fill="B8D0E8",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "A10:A18",
        fill="D9E7F4",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "B10:B18",
        fill="F7F1DA",
        font=Font(name="Calibri", size=11, bold=True, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    style_range(
        ws,
        "C10:G18",
        fill="F8FBFE",
        font=Font(name="Calibri", size=10, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )
    changes.append("stress-test block aligned to left module width")

    style_range(
        ws,
        "J2:M7",
        fill="F3F7FC",
        font=Font(name="Calibri", size=9, color=TEXT_DARK),
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
        border=thin_border(),
    )
    changes.append("right info panel moved away from main block and made readable")

    return changes
TIER_FREEBIE_MINUS = ["BUDGETS", "JAHR", "FIXKOSTEN", "NOTGROSCHEN", "PLANUNG"]
TIER_VOLLVERSION_PLUS = ["SCHULDEN", "MONATSABSCHLUSS", "STEUER", "SPARZIELE"]


def recreate_sheet(wb, title: str):
    if title in wb.sheetnames:
        idx = wb.sheetnames.index(title)
        existing = wb[title]
        wb.remove(existing)
        return wb.create_sheet(title=title, index=idx)
    return wb.create_sheet(title=title)


def apply_feature_sheet_frame(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    lead_lines: list[str],
    table_headers: list[str],
    table_rows: list[list[str]],
    notes_title: str,
    notes_lines: list[str],
    closing_lines: list[str],
) -> list[str]:
    changes: list[str] = []

    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    max_clear_row = max(ws.max_row, 32)
    max_clear_col = max(ws.max_column, 13)
    for row in ws.iter_rows(min_row=1, max_row=max_clear_row, min_col=1, max_col=max_clear_col):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name="Calibri", size=10, color=TEXT_DARK)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border()

    widths = {
        "A": 22, "B": 16, "C": 12, "D": 14,
        "E": 14, "F": 14, "G": 14, "H": 18,
        "I": 3, "J": 3, "K": 18, "L": 18, "M": 18,
    }
    for column_name, width in widths.items():
        ws.column_dimensions[column_name].width = width

    for idx in range(1, 33):
        ws.row_dimensions[idx].height = 22
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[5].height = 24

    ws.merge_cells("A1:H1")
    ws["A1"] = title
    style_range(
        ws,
        "A1:H1",
        fill=NAVY,
        font=Font(name="Calibri", size=18, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )

    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    style_range(
        ws,
        "A2:H2",
        fill=BLUE_LIGHT,
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )

    lead_row = 3
    for line in lead_lines:
        ws.merge_cells(f"A{lead_row}:H{lead_row}")
        ws[f"A{lead_row}"] = line
        style_range(
            ws,
            f"A{lead_row}:H{lead_row}",
            fill="F7FAFE",
            font=Font(name="Calibri", size=10, color=TEXT_DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=thin_border(),
        )
        lead_row += 1

    table_header_row = 5
    table_end_col = chr(ord("A") + len(table_headers) - 1)

    for col_idx, header in enumerate(table_headers, start=1):
        ws.cell(row=table_header_row, column=col_idx, value=header)
    style_range(
        ws,
        f"A{table_header_row}:{table_end_col}{table_header_row}",
        fill="D7E4F2",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin_border(),
    )

    current_row = table_header_row + 1
    for row_values in table_rows:
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=current_row, column=col_idx, value=value)
        style_range(
            ws,
            f"A{current_row}:{table_end_col}{current_row}",
            fill="FFFFFF",
            font=Font(name="Calibri", size=10, color=TEXT_DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=thin_border(),
        )
        current_row += 1

    notes_last_row = max(6, len(notes_lines) + 2)
    ws.merge_cells("K1:M1")
    ws["K1"] = notes_title
    style_range(
        ws,
        "K1:M1",
        fill=NAVY_DARK,
        font=Font(name="Calibri", size=10, bold=True, color=WHITE),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(NAVY_DARK),
    )

    for idx, line in enumerate(notes_lines, start=2):
        ws.merge_cells(f"K{idx}:M{idx}")
        ws[f"K{idx}"] = line
        style_range(
            ws,
            f"K{idx}:M{idx}",
            fill=BLUE_PALE,
            font=Font(name="Calibri", size=9, color=TEXT_DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=thin_border(),
        )
    for idx in range(len(notes_lines) + 2, notes_last_row + 1):
        ws.merge_cells(f"K{idx}:M{idx}")
        style_range(
            ws,
            f"K{idx}:M{idx}",
            fill=BLUE_PALE,
            font=Font(name="Calibri", size=9, color=TEXT_DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=thin_border(),
        )

    closing_header_row = current_row + 1
    ws.merge_cells(f"A{closing_header_row}:H{closing_header_row}")
    ws[f"A{closing_header_row}"] = "Praxis-Hinweise"
    style_range(
        ws,
        f"A{closing_header_row}:H{closing_header_row}",
        fill="EAF1FA",
        font=Font(name="Calibri", size=10, bold=True, color=NAVY_DARK),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin_border(),
    )

    footer_row = closing_header_row + 1
    for line in closing_lines:
        ws.merge_cells(f"A{footer_row}:H{footer_row}")
        ws[f"A{footer_row}"] = line
        style_range(
            ws,
            f"A{footer_row}:H{footer_row}",
            fill="FCFDFE",
            font=Font(name="Calibri", size=10, color=TEXT_DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=thin_border(),
        )
        footer_row += 1

    changes.append(f"{title} feature sheet scaffold applied")
    return changes


def apply_schulden_scaffold(ws: Worksheet) -> list[str]:
    return apply_feature_sheet_frame(
        ws,
        title="SCHULDEN",
        subtitle="Schuldenueberblick und Tilgungsplan fuer die Vollversion",
        lead_lines=[
            "Alle offenen Verbindlichkeiten gesammelt pflegen und monatlich aktualisieren.",
            "Mit Mindest-Rate und Ziel-Rate arbeiten, damit der Abbau planbar bleibt.",
        ],
        table_headers=["Position", "Restschuld", "Zins %", "Mindest-Rate", "Ziel-Rate", "Faellig", "Status"],
        table_rows=[
            ["Dispo", "0,00", "0,0", "0,00", "0,00", "TT.MM.JJJJ", "offen"],
            ["Kredit 1", "0,00", "0,0", "0,00", "0,00", "TT.MM.JJJJ", "offen"],
            ["Kredit 2", "0,00", "0,0", "0,00", "0,00", "TT.MM.JJJJ", "offen"],
            ["Ratenkauf", "0,00", "0,0", "0,00", "0,00", "TT.MM.JJJJ", "offen"],
            ["Privatdarlehen", "0,00", "0,0", "0,00", "0,00", "TT.MM.JJJJ", "offen"],
        ],
        notes_title="Schnellnutzung",
        notes_lines=[
            "1. Jede Schuld einzeln eintragen.",
            "2. Restschuld nur nach echtem Monatsabschluss aktualisieren.",
            "3. Ziel-Rate bewusst setzen und nicht still aendern.",
            "4. Sondertilgungen in Status/Notiz dokumentieren.",
        ],
        closing_lines=[
            "Prioritaet klar waehlen: kleine Schulden zuerst oder teuerste zuerst.",
            "Entscheidend ist ein ruhiger Plan, der dauerhaft tragfaehig bleibt.",
        ],
    )


def apply_monatsabschluss_scaffold(ws: Worksheet) -> list[str]:
    return apply_feature_sheet_frame(
        ws,
        title="MONATSABSCHLUSS",
        subtitle="Monatlicher Abschluss- und Kontrolllauf fuer den Haushalt",
        lead_lines=[
            "Dieses Blatt fuehrt den Monatsabschluss als kurze, wiederholbare Routine.",
            "Erst pruefen, dann schliessen, dann in den naechsten Monat gehen.",
        ],
        table_headers=["Pruefschritt", "Quelle", "Ergebnis", "Status", "Termin", "Owner", "Notiz"],
        table_rows=[
            ["Einnahmen vollstaendig", "MONAT", "offen", "zu pruefen", "Monatsende", "Haushalt", ""],
            ["Fixkosten geprueft", "FIXKOSTEN", "offen", "zu pruefen", "Monatsende", "Haushalt", ""],
            ["Variable Ausgaben geprueft", "HAUSHALTSBUCH", "offen", "zu pruefen", "Monatsende", "Haushalt", ""],
            ["Sparrate gebucht", "SPARZIELE", "offen", "zu pruefen", "Monatsende", "Haushalt", ""],
            ["Schuldenplan aktualisiert", "SCHULDEN", "offen", "zu pruefen", "Monatsende", "Haushalt", ""],
            ["Ruecklagen/Steuer geprueft", "STEUER", "offen", "zu pruefen", "Monatsende", "Haushalt", ""],
        ],
        notes_title="Abschlusslogik",
        notes_lines=[
            "1. Nur mit echten Buchungen arbeiten.",
            "2. Erst fehlende Daten nachtragen, dann abhaken.",
            "3. Status bewusst auf erledigt setzen.",
            "4. Notiz nur fuer echte Abweichungen nutzen.",
        ],
        closing_lines=[
            "Der Monatsabschluss ist die Kontrollschleife fuer alle anderen Vollversion-Module.",
            "Eine kurze, saubere Routine ist wertvoller als ein perfektes, nie gepflegtes Blatt.",
        ],
    )


def apply_steuer_scaffold(ws: Worksheet) -> list[str]:
    return apply_feature_sheet_frame(
        ws,
        title="STEUER",
        subtitle="Steuerbezogene Ruecklagen- und Kontrollflaeche fuer variable Einkommen",
        lead_lines=[
            "Dieses Blatt sammelt steuerrelevante Positionen und die geplante Ruecklage pro Monat.",
            "Parameter-Hinweise und Saetze bleiben die Eingabequelle, hier wird die Praxis geplant.",
        ],
        table_headers=["Bereich", "Basis", "Satz %", "Ruecklage/Monat", "Topf", "Termin", "Notiz"],
        table_rows=[
            ["Einkommensteuer", "0,00", "0,0", "0,00", "Steuer", "Monatsende", ""],
            ["Kirchensteuer", "0,00", "0,0", "0,00", "Steuer", "Monatsende", "0 = nicht relevant"],
            ["Solidaritaet", "0,00", "0,0", "0,00", "Steuer", "Monatsende", ""],
            ["Nebeneinkuenfte", "0,00", "0,0", "0,00", "Steuer", "Monatsende", ""],
            ["Kapitalertraege", "0,00", "0,0", "0,00", "Steuer", "Monatsende", ""],
        ],
        notes_title="Wichtige Regel",
        notes_lines=[
            "1. Nur relevante Bereiche pflegen.",
            "2. Saetze in PARAMETER kontrollieren, nicht hier erfinden.",
            "3. Ruecklage monatlich anpassen, wenn Basis sich aendert.",
            "4. Dieses Blatt ist Plan- und Kontrollflaeche, kein Steuerbescheid.",
        ],
        closing_lines=[
            "Ziel ist eine ruhige Steuer-Ruecklage statt spaeterer Ueberraschungen.",
            "Wer keine steuerrelevanten Nebenstroeme hat, kann das Blatt bewusst minimal halten.",
        ],
    )


def apply_sparziele_scaffold(ws: Worksheet) -> list[str]:
    return apply_feature_sheet_frame(
        ws,
        title="SPARZIELE",
        subtitle="Sparziele, Zielbetraege und Monatsraten geordnet an einem Ort",
        lead_lines=[
            "Ziele konkret beziffern und mit realistischen Monatsraten verbinden.",
            "Nur Ziele fuehren, die wirklich aktiv verfolgt werden.",
        ],
        table_headers=["Ziel", "Zielbetrag", "Aktuell", "Luecke", "Monatsrate", "Zielmonat", "Status"],
        table_rows=[
            ["Notgroschen", "0,00", "0,00", "0,00", "0,00", "JJJJ-MM", "aktiv"],
            ["Jahreskosten", "0,00", "0,00", "0,00", "0,00", "JJJJ-MM", "aktiv"],
            ["Urlaub", "0,00", "0,00", "0,00", "0,00", "JJJJ-MM", "optional"],
            ["Technik", "0,00", "0,00", "0,00", "0,00", "JJJJ-MM", "optional"],
            ["Puffer", "0,00", "0,00", "0,00", "0,00", "JJJJ-MM", "aktiv"],
        ],
        notes_title="Steuerung",
        notes_lines=[
            "1. Zielbetrag bewusst festlegen.",
            "2. Monatsrate nur aus echtem Ueberschuss planen.",
            "3. Status auf optional setzen, wenn das Ziel warten kann.",
            "4. Nach jedem Monatsabschluss Aktualisierung kurz nachziehen.",
        ],
        closing_lines=[
            "Ein Zielblatt funktioniert nur mit wenigen, klar priorisierten Zielen.",
            "Weniger Ziele, dafuer echte Fortschritte, ist die bessere Vollversion-Logik.",
        ],
    )


def ensure_feature_sheet(wb, title: str) -> list[str]:
    ws = recreate_sheet(wb, title)

    if title == "SCHULDEN":
        return apply_schulden_scaffold(ws)
    if title == "MONATSABSCHLUSS":
        return apply_monatsabschluss_scaffold(ws)
    if title == "STEUER":
        return apply_steuer_scaffold(ws)
    if title == "SPARZIELE":
        return apply_sparziele_scaffold(ws)

    raise ValueError(f"Unsupported feature sheet: {title}")


def apply_tier_transform(wb, tier: str) -> list[str]:
    changes: list[str] = []

    if tier == "PRO":
        changes.append("PRO tier selected; workbook kept as baseline")
        return changes

    if tier == "FREEBIE":
        for sheet_name in TIER_FREEBIE_MINUS:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                wb.remove(ws)
                changes.append(f"{sheet_name} removed for FREEBIE")
        return changes

    if tier == "VOLLVERSION":
        for sheet_name in TIER_VOLLVERSION_PLUS:
            changes.extend(ensure_feature_sheet(wb, sheet_name))
        return changes

    raise ValueError(f"Unsupported tier: {tier}")

def main() -> int:
    parser = argparse.ArgumentParser(description="Deterministic workbook builder for EGO XLSX.")
    parser.add_argument("--input", required=True, help="Path to input workbook")
    parser.add_argument("--output", required=True, help="Path to output workbook or snapshot json")
    parser.add_argument("--mode", choices=["apply", "snapshot"], required=True, help="Operation mode")
    parser.add_argument("--sheet", choices=["START", "HAUSHALTSBUCH", "MONAT", "BUDGETS", "FIXKOSTEN", "PLANUNG", "JAHR", "NOTGROSCHEN", "SCHULDEN", "MONATSABSCHLUSS", "STEUER", "SPARZIELE"], default="START", help="Target sheet")
    parser.add_argument("--tier", choices=["FREEBIE", "PRO", "VOLLVERSION"], default=None, help="Target workbook tier")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    wb = load_workbook(input_path)
    if args.tier:
        if args.mode == "snapshot":
            payload = {
                "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                "input": str(input_path),
                "tier": args.tier,
                "mode": args.mode,
                "sheetnames_before": list(wb.sheetnames),
            }
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"SNAPSHOT: {output_path}")
            return 0

        changes = apply_tier_transform(wb, args.tier)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        report = {
            "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "input": str(input_path),
            "output": str(output_path),
            "tier": args.tier,
            "mode": args.mode,
            "changes": changes,
            "sheetnames_after": list(wb.sheetnames),
        }
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0

    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {args.sheet}")

    ws = wb[args.sheet]

    if args.mode == "snapshot":
        payload = {
            "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "input": str(input_path),
            "sheet": args.sheet,
            "mode": args.mode,
            "snapshot": collect_snapshot(ws),
        }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"SNAPSHOT: {output_path}")
        return 0

    if args.sheet == "START":
        changes = apply_start_refine(ws)
    elif args.sheet == "HAUSHALTSBUCH":
        changes = apply_haushaltsbuch_scaffold(ws)
    elif args.sheet == "MONAT":
        changes = apply_monat_scaffold(ws)
    elif args.sheet == "BUDGETS":
        changes = apply_budgets_scaffold(ws)
    elif args.sheet == "FIXKOSTEN":
        changes = apply_fixkosten_scaffold(ws)
    elif args.sheet == "PLANUNG":
        changes = apply_planung_scaffold(ws)
    elif args.sheet == "JAHR":
        changes = apply_jahr_scaffold(ws)
    elif args.sheet == "NOTGROSCHEN":
        changes = apply_notgroschen_scaffold(ws)
    elif args.sheet == "SCHULDEN":
        changes = apply_schulden_scaffold(ws)
    elif args.sheet == "MONATSABSCHLUSS":
        changes = apply_monatsabschluss_scaffold(ws)
    elif args.sheet == "STEUER":
        changes = apply_steuer_scaffold(ws)
    else:
        changes = apply_sparziele_scaffold(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    report = {
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "input": str(input_path),
        "output": str(output_path),
        "sheet": args.sheet,
        "mode": args.mode,
        "changes": changes,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
