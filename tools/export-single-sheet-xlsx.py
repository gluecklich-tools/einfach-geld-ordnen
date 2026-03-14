from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()
    sheet_name = args.sheet

    if not input_path.exists():
        raise SystemExit(f"FAIL: input missing: {input_path}")

    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"FAIL: sheet missing: {sheet_name}")

    for name in list(wb.sheetnames):
        if name != sheet_name:
            wb.remove(wb[name])

    wb.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    verify_wb = load_workbook(output_path, read_only=True)
    output_sheet_names = list(verify_wb.sheetnames)
    verify_wb.close()

    payload = {
        "input": str(input_path),
        "output": str(output_path),
        "sheet": sheet_name,
        "output_sheet_count": len(output_sheet_names),
        "output_sheet_names": output_sheet_names,
        "output_sha256": hashlib.sha256(output_path.read_bytes()).hexdigest(),
    }
    print(json.dumps(payload, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())